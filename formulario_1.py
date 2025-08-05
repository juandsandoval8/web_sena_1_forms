from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import os, json, hashlib, tempfile
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = "clave_secreta_super_segura"

# Paths
BASE_PATH = "BaseDatosApp"
TEMP_PATH = os.path.join(BASE_PATH, "DatosTemporales")
CONSOLIDATED_PATH = os.path.join(BASE_PATH, "Consolidados")
CONFIG_PATH = os.path.join(BASE_PATH, "Config")
DATA_FILE = os.path.join(TEMP_PATH, "instructores_temp.json")
USERS_FILE = os.path.join(CONFIG_PATH, "usuarios.json")
TEMPLATE_FILE = os.path.join(BASE_PATH, "Plantilla_1.xlsx")  # Plantilla base

# ===================== UTILIDADES =====================
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def load_json(file_path):
    if os.path.exists(file_path):
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def save_json(file_path, data):
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def init_system():
    os.makedirs(TEMP_PATH, exist_ok=True)
    os.makedirs(CONSOLIDATED_PATH, exist_ok=True)
    os.makedirs(CONFIG_PATH, exist_ok=True)

    if not os.path.exists(USERS_FILE):
        default_users = {
            "instructor": {"password": hash_password("instructor123"), "role": "instructor"},
            "admin": {"password": hash_password("admin123"), "role": "administrador"},
            "dev": {"password": hash_password("dev123"), "role": "desarrollador"},
        }
        save_json(USERS_FILE, default_users)

    if not os.path.exists(DATA_FILE):
        save_json(DATA_FILE, [])

# ===================== LOGIN =====================
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        users = load_json(USERS_FILE)

        if username in users and users[username]["password"] == hash_password(password):
            session["user"] = username
            session["role"] = users[username]["role"]
            return redirect(url_for("dashboard"))
        else:
            flash("Usuario o contraseña incorrectos", "danger")
    return render_template("login.html")

@app.route("/dashboard")
def dashboard():
    if "role" not in session:
        return redirect(url_for("login"))

    role = session["role"]
    if role == "instructor":
        return redirect(url_for("instructor_form"))
    elif role == "administrador":
        return redirect(url_for("admin_panel"))
    elif role == "desarrollador":
        return redirect(url_for("dev_panel"))
    else:
        return redirect(url_for("login"))

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ===================== INSTRUCTOR =====================
@app.route("/instructor", methods=["GET", "POST"])
def instructor_form():
    if "role" not in session or session["role"] != "instructor":
        return redirect(url_for("login"))

    if request.method == "POST":
        form_data = dict(request.form)
        form_data["cumplimiento_adicionales"] = "cumplimiento_adicionales" in request.form
        form_data["cumple_perfil"] = "cumple_perfil" in request.form
        form_data["virtual"] = "virtual" in request.form   # ✅ Nuevo campo
        form_data["fecha_registro"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        registros = load_json(DATA_FILE) or []
        registros.append(form_data)
        save_json(DATA_FILE, registros)
        flash("Datos guardados correctamente", "success")
        return redirect(url_for("instructor_form"))

    return render_template("instructor.html", user=session["user"])

# ===================== ADMIN =====================
@app.route("/admin")
def admin_panel():
    if "role" not in session or session["role"] != "administrador":
        return redirect(url_for("login"))

    registros = load_json(DATA_FILE)

    # --- Filtros ---
    filtro_instructor = request.args.get("filtro_instructor", "").strip().lower()
    filtro_programa = request.args.get("filtro_programa", "").strip().lower()
    filtro_virtual = request.args.get("filtro_virtual", "")

    if filtro_instructor:
        registros = [r for r in registros if filtro_instructor in r.get("nombre_instructor", "").lower()]
    if filtro_programa:
        registros = [r for r in registros if filtro_programa in r.get("programa", "").lower()]
    if filtro_virtual == "true":
        registros = [r for r in registros if r.get("virtual", False)]
    elif filtro_virtual == "false":
        registros = [r for r in registros if not r.get("virtual", False)]

    return render_template("admin.html", registros=registros)

@app.route("/admin/detalle/<int:id>")
def admin_detalle(id):
    registros = load_json(DATA_FILE)
    if id < 0 or id >= len(registros):
        flash("Registro no encontrado", "danger")
        return redirect(url_for("admin_panel"))
    return render_template("detalle.html", registro=registros[id], id=id)

@app.route("/admin/eliminar/<int:id>")
def admin_eliminar(id):
    registros = load_json(DATA_FILE)
    if 0 <= id < len(registros):
        registros.pop(id)
        save_json(DATA_FILE, registros)
        flash("Registro eliminado correctamente", "success")
    return redirect(url_for("admin_panel"))

# ---- Plantilla Individual ----
@app.route("/admin/plantilla/<int:id>")
def admin_plantilla(id):
    registros = load_json(DATA_FILE)
    if id < 0 or id >= len(registros):
        flash("Registro no encontrado", "danger")
        return redirect(url_for("admin_panel"))

    if not os.path.exists(TEMPLATE_FILE):
        flash("No se encontró la plantilla Excel en el servidor", "danger")
        return redirect(url_for("admin_panel"))

    data = registros[id]
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    # ---- Cabecera ----
    ws["C4"] = data.get("regional", "")
    ws["C5"] = data.get("centro_formacion", "")
    ws["C6"] = data.get("lugar_desarrollo", "")
    ws["C7"] = data.get("programa", "")
    ws["C8"] = data.get("fecha_elaboracion", "")

    # ---- Fila dinámica ----
    fila_actual = 12
    mapeo_campos = {
        'B': 'competencia_programa',
        'C': 'tipo_competencia',
        'D': 'descripcion_perfil',
        'E': 'nombre_instructor',
        'F': 'tipo_vinculacion',
        'G': 'nivel_formacion',
        'H': 'titulo',
        'I': 'nivel_formacion2',
        'J': 'titulo2',
        'K': 'experiencia_tecnica',
        'L': 'experiencia_docente',
        'M': 'cumplimiento_adicionales',
        'N': 'observaciones'
    }
    for col, campo in mapeo_campos.items():
        valor = data.get(campo, "")
        if campo == "cumplimiento_adicionales":
            valor = "Sí" if valor else "No"
        ws[f"{col}{fila_actual}"] = valor

    if data.get("cumple_perfil", False):
        ws[f"O{fila_actual}"] = "SI (X)"
        ws[f"P{fila_actual}"] = "NO ( )"
    else:
        ws[f"O{fila_actual}"] = "SI ( )"
        ws[f"P{fila_actual}"] = "NO (X)"

    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp_file.name)
    tmp_file.close()

    return send_file(tmp_file.name, as_attachment=True, download_name=f"Instructor_{id+1}.xlsx")

# ---- Plantilla con Seleccionados ----
@app.route("/admin/plantilla_seleccion", methods=["POST"])
def admin_plantilla_seleccion():
    registros = load_json(DATA_FILE)
    seleccionados = request.form.getlist("seleccionados")

    if not seleccionados:
        flash("No seleccionaste ningún registro", "warning")
        return redirect(url_for("admin_panel"))

    if not os.path.exists(TEMPLATE_FILE):
        flash("No se encontró la plantilla Excel en el servidor", "danger")
        return redirect(url_for("admin_panel"))

    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    # Cabecera del primer seleccionado
    data_header = registros[int(seleccionados[0])]
    ws["C4"] = data_header.get("regional", "")
    ws["C5"] = data_header.get("centro_formacion", "")
    ws["C6"] = data_header.get("lugar_desarrollo", "")
    ws["C7"] = data_header.get("programa", "")
    ws["C8"] = data_header.get("fecha_elaboracion", "")

    fila_actual = 12
    for idx in seleccionados:
        data = registros[int(idx)]
        mapeo_campos = {
            'B': 'competencia_programa',
            'C': 'tipo_competencia',
            'D': 'descripcion_perfil',
            'E': 'nombre_instructor',
            'F': 'tipo_vinculacion',
            'G': 'nivel_formacion',
            'H': 'titulo',
            'I': 'nivel_formacion2',
            'J': 'titulo2',
            'K': 'experiencia_tecnica',
            'L': 'experiencia_docente',
            'M': 'cumplimiento_adicionales',
            'N': 'observaciones'
        }
        for col, campo in mapeo_campos.items():
            valor = data.get(campo, "")
            if campo == "cumplimiento_adicionales":
                valor = "Sí" if valor else "No"
            ws[f"{col}{fila_actual}"] = valor

        if data.get("cumple_perfil", False):
            ws[f"O{fila_actual}"] = "SI (X)"
            ws[f"P{fila_actual}"] = "NO ( )"
        else:
            ws[f"O{fila_actual}"] = "SI ( )"
            ws[f"P{fila_actual}"] = "NO (X)"

        fila_actual += 1

    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp_file.name)
    tmp_file.close()

    return send_file(tmp_file.name, as_attachment=True, download_name="Plantilla_Seleccionados.xlsx")

# ---- Plantilla con Todos ----
@app.route("/admin/plantilla_all")
def admin_plantilla_all():
    registros = load_json(DATA_FILE)
    if not registros:
        flash("No hay registros para generar en la plantilla", "warning")
        return redirect(url_for("admin_panel"))

    if not os.path.exists(TEMPLATE_FILE):
        flash("No se encontró la plantilla Excel en el servidor", "danger")
        return redirect(url_for("admin_panel"))

    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    # Cabecera del primer registro
    data_header = registros[0]
    ws["C4"] = data_header.get("regional", "")
    ws["C5"] = data_header.get("centro_formacion", "")
    ws["C6"] = data_header.get("lugar_desarrollo", "")
    ws["C7"] = data_header.get("programa", "")
    ws["C8"] = data_header.get("fecha_elaboracion", "")

    fila_actual = 12
    for data in registros:
        mapeo_campos = {
            'B': 'competencia_programa',
            'C': 'tipo_competencia',
            'D': 'descripcion_perfil',
            'E': 'nombre_instructor',
            'F': 'tipo_vinculacion',
            'G': 'nivel_formacion',
            'H': 'titulo',
            'I': 'nivel_formacion2',
            'J': 'titulo2',
            'K': 'experiencia_tecnica',
            'L': 'experiencia_docente',
            'M': 'cumplimiento_adicionales',
            'N': 'observaciones'
        }
        for col, campo in mapeo_campos.items():
            valor = data.get(campo, "")
            if campo == "cumplimiento_adicionales":
                valor = "Sí" if valor else "No"
            ws[f"{col}{fila_actual}"] = valor

        if data.get("cumple_perfil", False):
            ws[f"O{fila_actual}"] = "SI (X)"
            ws[f"P{fila_actual}"] = "NO ( )"
        else:
            ws[f"O{fila_actual}"] = "SI ( )"
            ws[f"P{fila_actual}"] = "NO (X)"

        fila_actual += 1

    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp_file.name)
    tmp_file.close()

    return send_file(tmp_file.name, as_attachment=True, download_name="Plantilla_Todos.xlsx")

@app.route("/admin/exportar")
def admin_exportar():
    registros = load_json(DATA_FILE)
    if not registros:
        flash("No hay registros para exportar", "warning")
        return redirect(url_for("admin_panel"))

    df = pd.DataFrame(registros)
    filename = os.path.join(CONSOLIDATED_PATH, f"instructores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    df.to_excel(filename, index=False)
    flash(f"Datos exportados a {filename}", "success")
    return redirect(url_for("admin_panel"))

@app.route("/admin/consolidar")
def admin_consolidar():
    registros = load_json(DATA_FILE)
    if not registros:
        flash("No hay registros para consolidar", "warning")
        return redirect(url_for("admin_panel"))

    folder = os.path.join(CONSOLIDATED_PATH, f"Consolidado_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    os.makedirs(folder, exist_ok=True)
    save_json(os.path.join(folder, "instructores.json"), registros)
    flash(f"Consolidado creado en {folder}", "success")
    return redirect(url_for("admin_panel"))

# ===================== DEV =====================
@app.route("/dev")
def dev_panel():
    if "role" not in session or session["role"] != "desarrollador":
        return redirect(url_for("login"))

    usuarios = load_json(USERS_FILE)
    registros = load_json(DATA_FILE)
    info = {
        "base": BASE_PATH,
        "users": len(usuarios),
        "registros": len(registros)
    }
    return render_template("developer.html", usuarios=usuarios, info=info)

@app.route("/dev/add_user", methods=["POST"])
def dev_add_user():
    usuarios = load_json(USERS_FILE)
    username = request.form["username"].strip()
    password = request.form["password"].strip()
    role = request.form["role"]

    if username in usuarios:
        flash("El usuario ya existe", "danger")
    else:
        usuarios[username] = {"password": hash_password(password), "role": role}
        save_json(USERS_FILE, usuarios)
        flash(f"Usuario {username} agregado", "success")
    return redirect(url_for("dev_panel"))

@app.route("/dev/delete/<username>")
def dev_delete_user(username):
    usuarios = load_json(USERS_FILE)
    if username in ["admin", "dev", "instructor"]:
        flash("No se pueden eliminar usuarios por defecto", "danger")
    elif username in usuarios:
        del usuarios[username]
        save_json(USERS_FILE, usuarios)
        flash(f"Usuario {username} eliminado", "success")
    return redirect(url_for("dev_panel"))

@app.route("/dev/verificar")
def dev_verificar():
    results = []
    for path_name, path in [
        ("Base", BASE_PATH),
        ("Temporales", TEMP_PATH),
        ("Consolidados", CONSOLIDATED_PATH),
        ("Config", CONFIG_PATH),
    ]:
        if os.path.exists(path):
            results.append(f"✓ Directorio {path_name}: OK")
        else:
            results.append(f"✗ Directorio {path_name}: FALTA")
    flash("\n".join(results), "info")
    return redirect(url_for("dev_panel"))

@app.route("/dev/backup")
def dev_backup():
    registros = load_json(DATA_FILE)
    usuarios = load_json(USERS_FILE)
    folder = os.path.join(CONSOLIDATED_PATH, f"Backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    os.makedirs(folder, exist_ok=True)
    save_json(os.path.join(folder, "usuarios.json"), usuarios)
    save_json(os.path.join(folder, "instructores.json"), registros)
    flash(f"Backup creado en {folder}", "success")
    return redirect(url_for("dev_panel"))

# ===================== MAIN =====================
if __name__ == "__main__":
    init_system()
    app.run(debug=True)
